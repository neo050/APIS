from datetime import datetime
import openpyxl as xl
import pywhatkit
from concurrent.futures import ThreadPoolExecutor
import os
import traceback
import time

# Constants for sheet column indices to improve readability
COL_PHONE = 2
COL_MSG = 3
COL_DATE = 4
COL_HOUR = 5
COL_MIN = 6
COL_LAST_DATE = 7


# Function to sanitize date string
def sanitize_date(date_str):
    return date_str.replace(" ", "").replace("-", "").replace("/","")


# Function to get current date and time as string
def get_current_datetime_str():
    now = datetime.now()
    return f"{now.day:02}-{now.month:02} {now.hour}:{now.minute}"


def process_workbook(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb.active
    check_date(sheet)
    safe_save_workbook(wb, file_name)


def check_date(sheet):
    today_date = sanitize_date(datetime.now().strftime("%d-%m-%Y"))

    for row in range(2, sheet.max_row + 1):
        date_cell_value = sanitize_date(str(sheet.cell(row, COL_DATE).value))
        print(date_cell_value,"   and   ",today_date)
        if today_date != date_cell_value:
            print("Not today")
            continue

        phone_number = f"+972{sheet.cell(row, COL_PHONE).value}"
        message = sheet.cell(row, COL_MSG).value
        hours = sheet.cell(row, COL_HOUR).value
        minutes = sheet.cell(row, COL_MIN).value
        last_date_cell = sheet.cell(row, COL_LAST_DATE)

        # Overwrite last_date_cell to current date and time
        last_date_cell.value = get_current_datetime_str()

        # Assuming you'll implement the logic to send the message
        # For example, call a function that handles scheduling or sending
        schedule_message(phone_number, message, hours, minutes)


def schedule_message(phone, message, hour, minute):
    # Implement the scheduling and error handling for sending messages
    try:
        pywhatkit.sendwhatmsg(phone, message, hour, minute)
    except Exception as e:
        print(f"Failed to send message: {e}")


def safe_save_workbook(workbook, file_name, attempts=5):
    for attempt in range(attempts):
        try:
            workbook.save(file_name)
            break
        except PermissionError as e:
            if attempt < attempts - 1:
                print(f"PermissionError on attempt {attempt + 1} to save '{file_name}'. Retrying...")
                time.sleep(1)  # Wait a second before retry
            else:
                raise e  # If final attempt fails, re-raise the exception


def main():
    executor = ThreadPoolExecutor()

    # Retrieve Excel files in the current directory, ignoring temporary files
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]

    futures = [executor.submit(process_workbook, file_name) for file_name in excel_files]

    # Wait for all threads to complete
    for future in futures:
        try:
            future.result()
        except Exception as e:
            print(f"Error processing a workbook: {e}")
            traceback.print_exc()

    print("Finished processing all workbooks.")


if __name__ == '__main__':
    print(f"Script started at {datetime.now()}")
    main()
